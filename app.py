import io
import streamlit as st
import pandas as pd
from datetime import date
from db import get_conn
import data_access
from pages import hisse_detay, anomali_backtest, master_analiz

st.set_page_config(
    page_title="Minerva",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@300;400;500&display=swap');

*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"] {
    font-family: 'IBM Plex Mono', monospace !important;
    background-color: #0c0c13 !important;
    color: #c0c0d0;
}
.stApp { background-color: #0c0c13 !important; }
.block-container {
    padding: 12px 20px 24px 20px !important;
    max-width: 100% !important;
}
section[data-testid="stSidebar"] { display: none !important; }
div[data-testid="stToolbar"]     { display: none !important; }
header[data-testid="stHeader"]   { display: none !important; }
footer { display: none !important; }

div[data-testid="stRadio"] > label,
div[data-testid="stRadio"] [data-testid="stWidgetLabel"],
div[data-testid="stRadio"] > div:first-child > p { display: none !important; }
div[data-testid="stRadio"] > div {
    display: flex !important; flex-direction: row !important;
    gap: 0 !important; flex-wrap: nowrap !important;
    border-bottom: 1px solid #1a1a24 !important;
    padding: 0 !important; margin: 0 !important;
    background: transparent !important;
}
div[data-testid="stRadio"] label {
    display: flex !important; align-items: center !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 11px !important; color: #3a3a55 !important;
    letter-spacing: 0.1em !important; text-transform: uppercase !important;
    padding: 9px 16px 10px !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -1px !important; cursor: pointer !important;
    white-space: nowrap !important; background: transparent !important;
    transition: color 0.1s !important;
}
div[data-testid="stRadio"] label:hover { color: #8888a8 !important; }
div[data-testid="stRadio"] label:has(input:checked) {
    color: #e0e0f0 !important;
    border-bottom: 2px solid #4d8ef0 !important;
}
div[data-testid="stRadio"] label > div:first-child { display: none !important; }
div[data-testid="stRadio"] label > p { margin: 0 !important; }

.stButton > button {
    background: transparent !important; border: 1px solid #1e1e30 !important;
    color: #4a4a68 !important; font-family: 'IBM Plex Mono', monospace !important;
    font-size: 10px !important; letter-spacing: 0.1em !important;
    text-transform: uppercase !important; border-radius: 2px !important;
    padding: 4px 12px !important;
}
.stButton > button:hover {
    border-color: #4d8ef0 !important; color: #4d8ef0 !important;
    background: #0d1a2e !important;
}
.stButton > button[kind="primary"] {
    background: #0d1a2e !important; border-color: #4d8ef0 !important;
    color: #4d8ef0 !important;
}
.stButton > button[kind="primary"]:hover { background: #152438 !important; }

[data-testid="stDownloadButton"] > button {
    background: transparent !important; border: 1px solid #2a4a2a !important;
    color: #4a8a4a !important; font-family: 'IBM Plex Mono', monospace !important;
    font-size: 10px !important; letter-spacing: 0.1em !important;
    text-transform: uppercase !important; border-radius: 2px !important;
    padding: 4px 12px !important; width: 100% !important;
}
[data-testid="stDownloadButton"] > button:hover {
    border-color: #22c55e !important; color: #22c55e !important;
    background: #0a1a0a !important;
}

div[data-baseweb="select"] > div {
    background: #0c0c13 !important; border: 1px solid #1e1e30 !important;
    border-radius: 2px !important; font-family: 'IBM Plex Mono', monospace !important;
    font-size: 11px !important; color: #8888a8 !important; min-height: 32px !important;
}
div[data-baseweb="select"] svg { color: #3a3a55 !important; }
[data-baseweb="popover"] { background: #12121e !important; border: 1px solid #1e1e30 !important; }
[role="option"] {
    background: #12121e !important; font-family: 'IBM Plex Mono', monospace !important;
    font-size: 11px !important; color: #8888a8 !important;
}
[role="option"]:hover { background: #1a1a2e !important; color: #e0e0f0 !important; }
[aria-selected="true"] { background: #1a1a2e !important; color: #4d8ef0 !important; }

input[type="date"], div[data-baseweb="input"] input {
    background: #0c0c13 !important; border: 1px solid #1e1e30 !important;
    border-radius: 2px !important; color: #8888a8 !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 11px !important; padding: 4px 8px !important;
}
textarea {
    background: #0c0c13 !important; border: 1px solid #1e1e30 !important;
    color: #c0c0d0 !important; font-family: 'IBM Plex Mono', monospace !important;
    font-size: 11px !important; border-radius: 2px !important;
}

.stTabs [data-baseweb="tab-list"] {
    background: transparent !important; border-bottom: 1px solid #1a1a24 !important; gap: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    font-family: 'IBM Plex Mono', monospace !important; font-size: 10px !important;
    letter-spacing: 0.1em !important; color: #3a3a55 !important;
    text-transform: uppercase !important; padding: 8px 14px !important;
    background: transparent !important;
}
.stTabs [aria-selected="true"] { color: #e0e0f0 !important; }
.stTabs [data-baseweb="tab-border"] { background: #4d8ef0 !important; }
.stTabs [data-baseweb="tab-panel"] { padding: 16px 0 0 !important; }

.stProgress > div > div { background: #4d8ef0 !important; border-radius: 0 !important; }
.stProgress > div { background: #1a1a24 !important; border-radius: 0 !important; height: 2px !important; }

.stDataFrame iframe { border: 1px solid #1a1a24 !important; }
[data-testid="stDataFrame"] div {
    font-family: 'IBM Plex Mono', monospace !important; font-size: 11px !important;
}

div[data-testid="stExpander"] {
    border: 1px solid #1a1a24 !important; border-radius: 2px !important;
    background: #0f0f18 !important;
}
div[data-testid="stExpander"] summary {
    font-family: 'IBM Plex Mono', monospace !important; font-size: 11px !important;
    color: #4a4a68 !important; padding: 8px 12px !important;
}

div[data-testid="stInfo"], div[data-testid="stWarning"], div[data-testid="stError"] {
    background: #0f0f18 !important; border-radius: 2px !important;
    font-family: 'IBM Plex Mono', monospace !important; font-size: 11px !important;
}


::-webkit-scrollbar { width: 3px; height: 3px; }
::-webkit-scrollbar-track { background: #0c0c13; }
::-webkit-scrollbar-thumb { background: #1e1e30; border-radius: 1px; }
::-webkit-scrollbar-thumb:hover { background: #2e2e44; }

div[data-testid="stVerticalBlock"] > div { padding-top: 0 !important; padding-bottom: 0 !important; }
div[data-testid="stVerticalBlock"] { gap: 0 !important; }
div[data-testid="stHorizontalBlock"] { gap: 12px !important; flex-wrap: nowrap !important; }
div[data-testid="column"] { padding: 0 !important; min-width: 0 !important; overflow: visible !important; }
div[data-testid="stVerticalBlockBorderWrapper"] { padding: 0 !important; }

/* Scroll container'ın kendi border'ını gizle */
div[data-testid="stVerticalBlockBorderWrapper"][style*="overflow"] {
    border: none !important;
    background: transparent !important;
}

/* ── Mobile ──────────────────────────────────────────────────────────────── */
@media (max-width: 768px) {
  .block-container { padding: 6px 8px 16px 8px !important; }

  /* Tüm kolon grupları dikey yığılsın */
  div[data-testid="stHorizontalBlock"] {
    flex-wrap: wrap !important;
    gap: 4px !important;
  }
  div[data-testid="column"] {
    min-width: 100% !important;
    flex: 1 1 100% !important;
  }

  /* Ok/boşluk ayırıcıları gizle */
  .mv-arrow { display: none !important; }
  /* Tarih etiketi gizle (tarih inputları kalır) */
  .mv-label { display: none !important; }

  /* Header: sağ taraftaki istatistikleri gizle, sol yeterli */
  .mv-header { flex-wrap: wrap !important; gap: 6px !important; }
  .mv-header-right { display: none !important; }

  /* Status bar: iki taraf üst üste */
  .mv-statusbar { flex-direction: column !important; gap: 3px !important; }

  /* Radio sekmeler daha dar */
  div[data-testid="stRadio"] label { padding: 8px 10px 9px !important; }

  /* Butonların font biraz daha büyük */
  .stButton > button { font-size: 11px !important; }
}
</style>
""", unsafe_allow_html=True)


@st.cache_data(ttl=300)
def _excel_tum_sikismalar() -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    conn = get_conn()
    df = data_access.sikisma_kayitlari_cek(conn)

    col_sirasi = [
        ("symbol",           "Hisse"),
        ("radar",            "Radar"),
        ("kutu_baslangic",   "Kutu Başlangıç"),
        ("kutu_bitis",       "Kutu Bitiş"),
        ("pencere_uzunlugu", "Pencere (Gün)"),
        ("kanal_yonu",       "Kanal Yönü"),
        ("cekirdek_zirve",   "Çekirdek Zirve"),
        ("cekirdek_dip",     "Çekirdek Dip"),
        ("fiziki_limit",     "Fiziki Limit"),
        ("efor_rasyosu",     "Efor Rasyosu"),
        ("sok_sayisi",       "Şok Sayısı"),
        ("sok_hacim_yuzdesi","Şok Hacim %"),
        ("m_norm",           "M-Norm"),
        ("trend_m",          "Trend Eğimi"),
        ("trend_c",          "Trend Sabiti"),
        ("kanal_ust_offset", "Kanal Üst Offset"),
        ("kanal_alt_offset", "Kanal Alt Offset"),
        ("olusturma_zaman",  "Oluşturma Zamanı"),
    ]
    mevcut  = [c for c, _ in col_sirasi if c in df.columns]
    isimler = {c: n for c, n in col_sirasi if c in df.columns}
    df_detay = (
        df[mevcut]
        .rename(columns=isimler)
        .sort_values(["Hisse", "Kutu Bitiş"], ascending=[True, False])
        .reset_index(drop=True)
    )

    # ── Özet sheet ──
    if not df.empty:
        ozet = (
            df.groupby("symbol")
            .agg(
                Radar1=("radar", lambda x: (x == "radar1").sum()),
                Radar2=("radar", lambda x: (x == "radar2").sum()),
                Toplam=("radar", "count"),
                Son_Sikisma=("kutu_bitis", "max"),
                En_Iyi_Efor=("efor_rasyosu", "max"),
                En_Cok_Sok=("sok_sayisi", "max"),
            )
            .reset_index()
            .rename(columns={
                "symbol":      "Hisse",
                "Radar1":      "Radar1 Sayısı",
                "Radar2":      "Radar2 Sayısı",
                "Toplam":      "Toplam Sıkışma",
                "Son_Sikisma": "Son Sıkışma",
                "En_Iyi_Efor": "En İyi Efor",
                "En_Cok_Sok":  "Max Şok",
            })
            .sort_values("En İyi Efor", ascending=False)
            .reset_index(drop=True)
        )
    else:
        ozet = pd.DataFrame()

    # ── Stil yardımcıları ──
    def stil_header(ws, renk_hex="0D1A2E", yazi_hex="4D8EF0"):
        dolgu = PatternFill(start_color=renk_hex, end_color=renk_hex, fill_type="solid")
        font  = Font(name="Calibri", bold=True, color=yazi_hex, size=10)
        for cell in ws[1]:
            cell.fill  = dolgu
            cell.font  = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    def otomatik_genislik(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 36)

    r1_dolgu  = PatternFill(start_color="080F1A", end_color="080F1A", fill_type="solid")
    r2_dolgu  = PatternFill(start_color="160F04", end_color="160F04", fill_type="solid")
    def_dolgu = PatternFill(start_color="0C0C13", end_color="0C0C13", fill_type="solid")
    veri_font = Font(name="Calibri", color="C0C0D0", size=9)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Detay sheet
        df_detay.to_excel(writer, sheet_name="Sıkışma Kayıtları", index=False)
        ws_d = writer.sheets["Sıkışma Kayıtları"]
        stil_header(ws_d)

        radar_idx = next(
            (i for i, cell in enumerate(ws_d[1], 1) if cell.value == "Radar"), None
        )
        for row in ws_d.iter_rows(min_row=2):
            rv    = row[radar_idx - 1].value if radar_idx else None
            dolgu = r1_dolgu if rv == "radar1" else (r2_dolgu if rv == "radar2" else def_dolgu)
            for cell in row:
                cell.fill      = dolgu
                cell.font      = veri_font
                cell.alignment = Alignment(vertical="center")
                ws_d.row_dimensions[cell.row].height = 15
        otomatik_genislik(ws_d)

        # Özet sheet
        if not ozet.empty:
            ozet.to_excel(writer, sheet_name="Hisse Özeti", index=False)
            ws_o = writer.sheets["Hisse Özeti"]
            stil_header(ws_o, renk_hex="0A1A0A", yazi_hex="22C55E")
            for row in ws_o.iter_rows(min_row=2):
                for cell in row:
                    cell.fill      = def_dolgu
                    cell.font      = veri_font
                    cell.alignment = Alignment(vertical="center")
                    ws_o.row_dimensions[cell.row].height = 15
            otomatik_genislik(ws_o)

    buf.seek(0)
    return buf.read()


@st.cache_data(ttl=300)
def _anomali_stats():
    conn = get_conn()
    return pd.read_sql("""
        SELECT
            (SELECT COUNT(*) FROM stocks WHERE is_active = true) AS hisse_sayisi,
            COUNT(*)                                              AS toplam,
            MAX(baslangic_zaman)                                  AS son_guncelleme
        FROM anomali_kayitlari
    """, conn).iloc[0]


@st.cache_data(ttl=300)
def _hisse_listesi():
    conn = get_conn()
    return pd.read_sql(
        "SELECT symbol FROM stocks WHERE is_active = true ORDER BY symbol", conn
    )["symbol"].tolist()


try:
    hisseler = _hisse_listesi()
except Exception:
    hisseler = []

try:
    sa = _anomali_stats()
    a_hisse   = int(sa["hisse_sayisi"])
    a_toplam  = int(sa["toplam"])
    a_son     = sa["son_guncelleme"]
    a_son_str = a_son.strftime("%Y-%m-%d %H:%M") if a_son is not None else "—"
except Exception:
    a_hisse = a_toplam = 0
    a_son_str = "—"

# ── Header ──
st.markdown(f"""
<div class="mv-header" style="display:flex;align-items:center;justify-content:space-between;padding:10px 0 0 0">
  <div style="display:flex;align-items:center;gap:10px">
    <div style="width:22px;height:22px;border-radius:50%;background:#4d8ef0;
                display:flex;align-items:center;justify-content:center;
                font-size:11px;font-weight:500;color:#fff;flex-shrink:0">M</div>
    <span style="font-size:13px;font-weight:500;color:#e0e0f0;letter-spacing:0.04em">Minerva</span>
    <span style="color:#1e1e30;font-size:13px">·</span>
    <span style="font-size:11px;color:#3a3a55;letter-spacing:0.06em">ANOMALİ + SIKIŞMA</span>
  </div>
  <div class="mv-header-right" style="display:flex;align-items:center;gap:16px">
    <span style="font-size:10px;color:#3a3a55;letter-spacing:0.06em">
      piyasa <span style="color:#6a6a88">BIST</span>
    </span>
    <span style="font-size:10px;color:#3a3a55">·</span>
    <span style="font-size:10px;color:#3a3a55;letter-spacing:0.06em">
      son güncelleme <span style="color:#6a6a88">{a_son_str}</span>
    </span>
    <span style="font-size:10px;color:#3a3a55">·</span>
    <span style="font-size:10px;color:#22c55e;letter-spacing:0.06em">● ready</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Nav + Hisse seçici + Export ──
col_nav, col_hisse, col_export = st.columns([7, 2, 1.4])

with col_nav:
    sayfa = st.radio(
        "nav",
        ["Genel Bakis", "Hisse Detay", "Backtest"],
        horizontal=True,
        label_visibility="collapsed",
        key="nav",
    )

with col_hisse:
    if sayfa in ("Genel Bakis", "Hisse Detay"):
        # Hisse Detay manages its own selector inside the page
        secilen = st.session_state.get("hd_hisse", "")
    else:
        secilen = st.selectbox(
            "hisse", hisseler,
            label_visibility="collapsed",
            key="nav_hisse",
        )

with col_export:
    try:
        excel_bytes = _excel_tum_sikismalar()
        dosya_adi = f"minerva_sikismalar_{date.today().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="↓ xlsx",
            data=excel_bytes,
            file_name=dosya_adi,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Tüm hisselerin geçmiş sıkışma verilerini Excel olarak indir",
        )
    except Exception as e:
        st.button("↓ xlsx", disabled=True, use_container_width=True, help=str(e))

# ── Status bar ──
ticker_html = f'ticker — {secilen}' if secilen else 'tıkla → aç'
st.markdown(f"""
<div class="mv-statusbar" style="font-size:10px;color:#2e2e48;letter-spacing:0.06em;
            padding:5px 0 12px 0;border-bottom:1px solid #12121e;
            display:flex;justify-content:space-between">
  <span>
    takip <span style="color:#4a4a68">{a_hisse}</span> hisse
    &nbsp;·&nbsp; toplam <span style="color:#4a4a68">{a_toplam:,}</span> anomali
  </span>
  <span style="color:#2e2e48">{ticker_html}</span>
</div>
""", unsafe_allow_html=True)

# ── Page routing ──
if sayfa == "Genel Bakis":
    master_analiz.goster()
elif sayfa == "Hisse Detay":
    hisse_detay.goster(hisseler)
elif sayfa == "Backtest":
    anomali_backtest.goster()

# ── Footer ──
st.markdown(f"""
<div style="margin-top:40px;padding:8px 0;border-top:1px solid #12121e;
            display:flex;justify-content:space-between;align-items:center">
  <span style="font-size:10px;color:#2a2a40;letter-spacing:0.06em">
    <span style="color:#22c55e">●</span>
    &nbsp;db · synced
  </span>
  <span style="font-size:10px;color:#2a2a40;letter-spacing:0.06em">Minerva</span>
</div>
""", unsafe_allow_html=True)
